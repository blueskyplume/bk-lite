from io import BytesIO

from django.http import HttpResponse
from django_filters import rest_framework as filters
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from rest_framework import permissions
from rest_framework.decorators import action
from rest_framework.response import Response

from apps.core.utils.viewset_utils import LanguageViewSet
from apps.system_mgmt.models import UserLoginLog
from apps.system_mgmt.serializers.user_login_log_serializer import UserLoginLogSerializer


class UserLoginLogFilter(filters.FilterSet):
    """用户登录日志过滤器"""

    # 登录状态筛选：success/failed
    status = filters.CharFilter(field_name="status", lookup_expr="exact")

    # 用户名筛选：支持精确匹配和模糊匹配
    username = filters.CharFilter(field_name="username", lookup_expr="icontains")

    # 登录时间范围筛选
    login_time_start = filters.DateTimeFilter(field_name="created_at", lookup_expr="gte")
    login_time_end = filters.DateTimeFilter(field_name="created_at", lookup_expr="lte")

    # 源IP筛选
    source_ip = filters.CharFilter(field_name="source_ip", lookup_expr="icontains")

    # 地理位置筛选
    location = filters.CharFilter(field_name="location", lookup_expr="icontains")

    # 操作系统筛选
    os_info = filters.CharFilter(field_name="os_info", lookup_expr="icontains")

    # 浏览器筛选
    browser_info = filters.CharFilter(field_name="browser_info", lookup_expr="icontains")

    class Meta:
        model = UserLoginLog
        fields = [
            "status",
            "username",
            "login_time_start",
            "login_time_end",
            "source_ip",
            "location",
            "os_info",
            "browser_info",
        ]


class UserLoginLogViewSet(LanguageViewSet):
    """
    用户登录日志视图集

    提供登录日志的查询和筛选功能，不支持创建、修改、删除操作
    登录日志由系统自动记录
    """

    queryset = UserLoginLog.objects.all().order_by("-created_at")
    serializer_class = UserLoginLogSerializer
    filterset_class = UserLoginLogFilter
    permission_classes = [permissions.IsAuthenticated]

    def get_http_method_names(self):
        """动态返回允许的HTTP方法"""
        # export_excel action 允许 POST，其他只允许 GET
        if self.action == "export_excel":
            return ["post", "options"]
        return ["get", "head", "options"]

    def http_method_not_allowed(self, request, *args, **kwargs):
        """检查HTTP方法是否允许"""
        if request.method.lower() not in self.get_http_method_names():
            return super().http_method_not_allowed(request, *args, **kwargs)
        return super().http_method_not_allowed(request, *args, **kwargs)

    @action(detail=False, methods=["get"])
    def statistics(self, request):
        """
        获取登录日志统计信息

        返回：
        - total: 总登录次数
        - success_count: 成功登录次数
        - failed_count: 失败登录次数
        - success_rate: 成功率
        """
        # 应用过滤器
        queryset = self.filter_queryset(self.get_queryset())

        total = queryset.count()
        success_count = queryset.filter(status=UserLoginLog.STATUS_SUCCESS).count()
        failed_count = queryset.filter(status=UserLoginLog.STATUS_FAILED).count()

        success_rate = round(success_count / total * 100, 2) if total > 0 else 0

        return Response(
            {
                "total": total,
                "success_count": success_count,
                "failed_count": failed_count,
                "success_rate": success_rate,
            }
        )

    @action(detail=False, methods=["post"])
    def export_excel(self, request):
        """
        导出登录日志为Excel文件

        支持两种导出模式：
        1. 筛选导出：传入筛选条件（与list接口相同）
        2. 指定ID导出：传入 selected_ids 列表

        筛选条件参数：
        - status: 登录状态
        - username: 用户名
        - login_time_start: 开始时间
        - login_time_end: 结束时间
        - source_ip: 源IP
        - location: 地理位置
        - os_info: 操作系统
        - browser_info: 浏览器

        指定ID参数：
        - selected_ids: ID列表，如 [1, 2, 3]

        返回：Excel文件流
        """
        # 从POST请求体中获取参数
        selected_ids = request.data.get("selected_ids", [])

        # 构建queryset
        queryset = self.get_queryset()

        # 如果指定了 selected_ids，则直接按ID筛选
        if selected_ids:
            queryset = queryset.filter(id__in=selected_ids)
        else:
            # 否则应用筛选条件
            filter_params = request.data
            filterset = self.filterset_class(filter_params, queryset=queryset)
            if filterset.is_valid():
                queryset = filterset.qs
            else:
                # 如果筛选参数无效，返回错误
                return Response({"result": False, "message": "Invalid filter parameters", "errors": filterset.errors}, status=400)

        # 创建Excel工作簿
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "用户登录日志"

        # 设置表头
        headers = [
            "用户名",
            "域名",
            "登录时间",
            "源IP地址",
            "地理位置",
            "浏览器",
            "操作系统",
            "登录状态",
            "失败原因",
        ]

        # 写入表头
        sheet.append(headers)

        # 设置表头样式
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_num, _ in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # 设置列宽
        column_widths = [20, 20, 22, 18, 25, 20, 20, 12, 40]
        for i, width in enumerate(column_widths, 1):
            sheet.column_dimensions[sheet.cell(row=1, column=i).column_letter].width = width

        # 获取状态显示翻译
        status_display_map = {
            UserLoginLog.STATUS_SUCCESS: self.loader.get("status.success") or "成功",
            UserLoginLog.STATUS_FAILED: self.loader.get("status.failed") or "失败",
        }

        # 写入数据
        for log in queryset:
            row_data = [
                log.username,
                log.domain,
                log.created_at.strftime("%Y-%m-%d %H:%M:%S") if log.created_at else "-",
                log.source_ip,
                log.location or "-",
                log.browser_info or "-",
                log.os_info or "-",
                status_display_map.get(log.status, log.status),
                log.failure_reason or "-",
            ]
            sheet.append(row_data)

            # 设置数据行样式
            row_num = sheet.max_row
            for col_num in range(1, len(headers) + 1):
                cell = sheet.cell(row=row_num, column=col_num)
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # 保存到BytesIO
        file_stream = BytesIO()
        workbook.save(file_stream)
        file_stream.seek(0)

        # 生成文件名
        from datetime import datetime

        filename = f"用户登录日志_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        # 返回Excel文件
        response = HttpResponse(file_stream.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        return response
