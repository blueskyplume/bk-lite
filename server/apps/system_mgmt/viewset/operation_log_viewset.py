from io import BytesIO

from django.http import HttpResponse
from django_filters import rest_framework as filters
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from rest_framework import permissions
from rest_framework.decorators import action
from rest_framework.response import Response

from apps.core.utils.viewset_utils import LanguageViewSet
from apps.system_mgmt.models import OperationLog
from apps.system_mgmt.serializers.operation_log_serializer import OperationLogSerializer


class OperationLogFilter(filters.FilterSet):
    """操作日志过滤器"""

    # 操作时间范围筛选
    operation_time_start = filters.DateTimeFilter(field_name="created_at", lookup_expr="gte")
    operation_time_end = filters.DateTimeFilter(field_name="created_at", lookup_expr="lte")

    # 应用筛选（多选）
    app = filters.CharFilter(field_name="app", lookup_expr="in", method="filter_app_in")

    # 用户名筛选（模糊）
    username = filters.CharFilter(field_name="username", lookup_expr="icontains")

    # 操作类型筛选（精准）
    action_type = filters.CharFilter(field_name="action_type", lookup_expr="exact")

    # 概要描述筛选（模糊）
    summary = filters.CharFilter(field_name="summary", lookup_expr="icontains")

    class Meta:
        model = OperationLog
        fields = [
            "operation_time_start",
            "operation_time_end",
            "app",
            "username",
            "action_type",
            "summary",
        ]

    def filter_app_in(self, queryset, name, value):
        """应用多选筛选"""
        if not value:
            return queryset
        # 支持逗号分隔的多个应用
        app_list = [app.strip() for app in value.split(",") if app.strip()]
        if app_list:
            return queryset.filter(app__in=app_list)
        return queryset


class OperationLogViewSet(LanguageViewSet):
    """
    操作日志视图集

    提供操作日志的查询和筛选功能，不支持创建、修改、删除操作
    操作日志由系统自动记录
    """

    queryset = OperationLog.objects.all().order_by("-created_at")
    serializer_class = OperationLogSerializer
    filterset_class = OperationLogFilter
    permission_classes = [permissions.IsAuthenticated]

    def get_http_method_names(self):
        """动态返回允许的HTTP方法"""
        # export_excel action 允许 POST，其他只允许 GET
        if self.action == "export_excel":
            return ["post", "options"]
        return ["get", "head", "options"]

    def http_method_not_allowed(self, request, *args, **kwargs):
        """检查HTTP方法是否允许"""
        if request.method.lower() not in self.get_http_method_names():
            return super().http_method_not_allowed(request, *args, **kwargs)
        return super().http_method_not_allowed(request, *args, **kwargs)

    def retrieve(self, request, *args, **kwargs):
        """禁用详情接口"""
        return Response({"result": False, "message": "Detail view is not supported"}, status=405)

    @action(detail=False, methods=["post"])
    def export_excel(self, request):
        """
        导出操作日志为Excel文件

        支持两种导出模式：
        1. 筛选导出：传入筛选条件（与list接口相同）
        2. 指定ID导出：传入 selected_ids 列表

        筛选条件参数：
        - operation_time_start: 开始时间
        - operation_time_end: 结束时间
        - app: 应用（多选，逗号分隔，如 "monitor,cmdb,node"）
        - username: 用户名（模糊匹配）
        - action_type: 操作类型（精准匹配：create/update/delete/execute）
        - summary: 概要描述（模糊匹配）

        指定ID参数：
        - selected_ids: ID列表，如 [1, 2, 3]

        返回：Excel文件流
        """
        # 从POST请求体中获取参数
        selected_ids = request.data.get("selected_ids", [])

        # 构建queryset
        queryset = self.get_queryset()

        # 如果指定了 selected_ids，则直接按ID筛选
        if selected_ids:
            queryset = queryset.filter(id__in=selected_ids)
        else:
            # 否则应用筛选条件
            filter_params = request.data
            filterset = self.filterset_class(filter_params, queryset=queryset)
            if filterset.is_valid():
                queryset = filterset.qs
            else:
                # 如果筛选参数无效，返回错误
                return Response({"result": False, "message": "Invalid filter parameters", "errors": filterset.errors}, status=400)

        # 创建Excel工作簿
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "用户操作日志"

        # 设置表头
        headers = [
            "用户名",
            "域名",
            "操作时间",
            "源IP地址",
            "应用",
            "操作类型",
            "概要描述",
        ]

        # 写入表头
        sheet.append(headers)

        # 设置表头样式
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_num, _ in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # 设置列宽
        column_widths = [20, 20, 22, 18, 25, 15, 50]
        for i, width in enumerate(column_widths, 1):
            sheet.column_dimensions[sheet.cell(row=1, column=i).column_letter].width = width

        # 获取操作类型显示翻译
        action_type_display_map = {
            OperationLog.ACTION_CREATE: self.loader.get("action_type.create") or "创建",
            OperationLog.ACTION_UPDATE: self.loader.get("action_type.update") or "更新",
            OperationLog.ACTION_DELETE: self.loader.get("action_type.delete") or "删除",
            OperationLog.ACTION_EXECUTE: self.loader.get("action_type.execute") or "执行",
        }

        # 写入数据
        for log in queryset:
            row_data = [
                log.username,
                log.domain,
                log.created_at.strftime("%Y-%m-%d %H:%M:%S") if log.created_at else "-",
                log.source_ip,
                log.app or "-",
                action_type_display_map.get(log.action_type, log.action_type),
                log.summary or "-",
            ]
            sheet.append(row_data)

            # 设置数据行样式
            row_num = sheet.max_row
            for col_num in range(1, len(headers) + 1):
                cell = sheet.cell(row=row_num, column=col_num)
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # 保存到BytesIO
        file_stream = BytesIO()
        workbook.save(file_stream)
        file_stream.seek(0)

        # 生成文件名
        from datetime import datetime

        filename = f"用户操作日志_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        # 返回Excel文件
        response = HttpResponse(file_stream.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        return response
